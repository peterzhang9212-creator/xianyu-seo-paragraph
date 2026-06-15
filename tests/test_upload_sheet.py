from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.upload_sheet import generate_upload_sheet


def _write_content_file(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["导出商品名", "标题+正文+标签", "发货属地"])
    sheet.append(["杭州教师SEO资料", "标题A\n正文A\n#标签", "杭州"])
    workbook.save(path)


def test_generate_upload_sheet_matches_folder_names(tmp_path: Path) -> None:
    content_path = tmp_path / "content.xlsx"
    _write_content_file(content_path)
    material_root = tmp_path / "materials"
    (material_root / "杭州教师SEO资料").mkdir(parents=True)
    output_path = tmp_path / "upload.xlsx"

    result = generate_upload_sheet(
        content_path=content_path,
        material_root=material_root,
        output_path=output_path,
    )

    saved = load_workbook(result.output_path)
    try:
        sheet = saved.active
        assert sheet.cell(row=4, column=1).value == "标题A\n正文A\n#标签"
        assert sheet.cell(row=4, column=2).value == str(material_root / "杭州教师SEO资料")
        assert sheet.cell(row=4, column=3).value == "杭州"
    finally:
        saved.close()
