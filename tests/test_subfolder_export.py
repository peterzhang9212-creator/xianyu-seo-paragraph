from pathlib import Path

from openpyxl import load_workbook

from core.subfolder_export import collect_subfolder_names, export_subfolder_names_to_excel


def test_collect_subfolder_names_returns_sorted_first_level_names(tmp_path: Path) -> None:
    (tmp_path / "b_folder").mkdir()
    (tmp_path / "a_folder").mkdir()
    (tmp_path / "a_folder" / "nested").mkdir()
    (tmp_path / "note.txt").write_text("x", encoding="utf-8")

    assert collect_subfolder_names(tmp_path) == ["a_folder", "b_folder"]


def test_export_subfolder_names_to_excel_writes_single_column(tmp_path: Path) -> None:
    output_path = tmp_path / "subfolders.xlsx"

    export_subfolder_names_to_excel(["商品A", "商品B"], output_path)

    workbook = load_workbook(output_path)
    try:
        sheet = workbook.active
        assert sheet.cell(row=1, column=1).value == "子文件夹名称"
        assert sheet.cell(row=2, column=1).value == "商品A"
        assert sheet.cell(row=3, column=1).value == "商品B"
    finally:
        workbook.close()
