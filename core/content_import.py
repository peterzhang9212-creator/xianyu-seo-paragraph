from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .models import ContentRow


REQUIRED_COLUMNS = ("导出商品名", "标题+正文+标签", "发货属地")


def load_content_rows(path: str | Path) -> list[ContentRow]:
    workbook = load_workbook(Path(path), read_only=False, data_only=True)
    try:
        sheet = workbook.active
        header_cells = next(sheet.iter_rows(max_row=1, values_only=True), None)
        if not header_cells:
            raise ValueError("正文文件没有表头")

        headers = [str(value).strip() if value is not None else "" for value in header_cells]
        missing = [column for column in REQUIRED_COLUMNS if column not in headers]
        if missing:
            raise ValueError(f"正文文件缺少必填列: {', '.join(missing)}")

        index_map = {column: headers.index(column) for column in REQUIRED_COLUMNS}
        rows: list[ContentRow] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not values or all(value in (None, "") for value in values):
                continue
            rows.append(
                ContentRow(
                    product_name=str(values[index_map["导出商品名"]] or "").strip(),
                    content=str(values[index_map["标题+正文+标签"]] or "").strip(),
                    location=str(values[index_map["发货属地"]] or "").strip(),
                )
            )
        if not rows:
            raise ValueError("正文文件没有可用数据")
        return rows
    finally:
        workbook.close()
